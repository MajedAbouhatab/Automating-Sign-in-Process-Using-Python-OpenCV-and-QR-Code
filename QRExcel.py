# Use the folllowing command to deploy in one exe file
# pyinstaller --clean --onefile --noconsole --add-data '*.wav;.' .\QRExcel.py

import ctypes
from os.path import join, exists, dirname, abspath, isfile

# If we have QR then we need QR Codes 
if exists('QR'):
    import qrcode
    from PIL import Image
    try:
        for line in open('QR.txt', 'r'):
            qrcode.make(line.strip(),border=1).save('QR\\' + line.strip() + '.png')
    # General error
    except Exception as e:
        pass # ctypes.windll.user32.MessageBoxW(None, str(e), 'Error', 0)
else:
    from datetime import datetime
    from openpyxl import load_workbook, Workbook
    from winsound import PlaySound, SND_LOOP, SND_ASYNC
    from cv2 import VideoCapture, QRCodeDetector, waitKey, imshow, flip

    def SoundEffect(f,s):
        PlaySound(join(dirname(abspath(__file__)), f), s)

    def ExcelIsOpen(e):
        SoundEffect('LockedFile.wav', SND_LOOP)
        ctypes.windll.user32.MessageBoxW(None, str(e) + '\n Please close ' + FileName + ' and try to scan again', 'Excel file still open', 0)

    FileName='QRExcel.xlsx'
    SoundEffect('Start.wav', SND_ASYNC)

    # Start Video Capture
    VC = VideoCapture(0)

    # Keep Video Capture open until Esc
    while waitKey(1) != 27:
        try:
            AttendeeName = QRCodeDetector().detectAndDecode(VC.read()[1])[0]
            if AttendeeName:

                # Create Excel file if it does not exist
                if not isfile(FileName):
                    Workbook().save(FileName)
                workbook = load_workbook(FileName)
                Sheet = workbook.active

                # Find Today's column
                AttendanceColumn = 0
                for Cell in Sheet[1][1:]:
                    if Cell.value and str(datetime.today())[:10] == Cell.value:
                        AttendanceColumn = Cell.column
                        break

                # Create Today's column
                if not AttendanceColumn:
                    AttendanceColumn = len(Sheet[1]) + 1
                    Sheet.cell(1,AttendanceColumn).value = str(datetime.today())[:10]
                    try:
                        workbook.save(FileName)
                    except Exception as e:
                        ExcelIsOpen(e)
                        continue

                # List of current Attendees        
                Attendees = [i[0] for i in Sheet.values][1:]

                # Add Attendee if new
                if AttendeeName not in Attendees:
                    Attendees.append(AttendeeName)
                    Sheet.cell(len(Attendees) + 1, 1).value = AttendeeName

                # Put timestamp only if the cell is blank
                if Sheet.cell(Attendees.index(AttendeeName) + 2, AttendanceColumn).value is None:
                    Sheet.cell(Attendees.index(AttendeeName) + 2, AttendanceColumn).value = datetime.now().strftime('%H:%M:%S')
                    Sheet.views.sheetView[0].selection[0].sqref = 'A1'
                    try:
                        workbook.save(FileName)
                        SoundEffect('Success.wav', SND_LOOP)
                    except Exception as e:
                        ExcelIsOpen(e)
                else:
                    SoundEffect('Duplicate.wav', SND_LOOP)

            # Show what computer sees
            imshow('ACE QR Code Scanner - Press Esc to close', flip(VC.read()[1], 1))

        # General error in the loop
        except Exception as e:
            pass # ctypes.windll.user32.MessageBoxW(None, str(e), 'Error', 0)
    
    # Exiting gracefully
    SoundEffect('End.wav', SND_LOOP)
